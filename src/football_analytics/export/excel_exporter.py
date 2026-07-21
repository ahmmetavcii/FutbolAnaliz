"""Excel report exporter: 18 fixed sheets with formatting and validation."""

from __future__ import annotations

import os
import tempfile
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

EXCEL_SHEET_NAMES: tuple[str, ...] = (
    "Match Summary",
    "Player Summary",
    "Goalkeeper Summary",
    "Team Summary",
    "Visibility Quality",
    "Camera Coverage",
    "Jersey Results",
    "Global Identity Mapping",
    "Identity Consistency",
    "Chunk Status",
    "Errors and Warnings",
    "Configuration",
    "Match Events",
    "Goals and Assists",
    "Shots",
    "Substitutions",
    "Officials",
    "Manual Corrections",
)

_HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_QUALITY_GOOD = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_QUALITY_WARN = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_QUALITY_BAD = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Columns holding scores in [0, 1] that should be traffic-light highlighted.
_QUALITY_TOKENS = ("confidence", "quality", "consistency", "coverage", "visibility", "score")
_PERCENT_TOKENS = ("_pct", "percent", "_ratio", "_rate", "share")
_INTEGER_TOKENS = ("frame", "count", "chunk_id", "minute", "jersey_number", "_id_num")
_QUALITY_WARN_THRESHOLD = 0.75
_QUALITY_BAD_THRESHOLD = 0.5


def _number_format_for(column_name: str, series: pd.Series) -> str | None:
    lowered = column_name.lower()
    if any(token in lowered for token in _PERCENT_TOKENS):
        return "0.0%"
    if any(token in lowered for token in _QUALITY_TOKENS):
        return "0.000"
    if pd.api.types.is_integer_dtype(series) or any(
        token in lowered for token in _INTEGER_TOKENS
    ):
        return "#,##0"
    if pd.api.types.is_float_dtype(series):
        return "0.00"
    return None


def _is_quality_column(column_name: str, series: pd.Series) -> bool:
    lowered = column_name.lower()
    if not any(token in lowered for token in _QUALITY_TOKENS):
        return False
    numeric = pd.to_numeric(series, errors="coerce").dropna()
    if numeric.empty:
        return True
    return bool((numeric >= 0.0).all() and (numeric <= 1.0).all())


def _quality_fill(value: Any) -> PatternFill | None:
    try:
        score = float(value)
    except (TypeError, ValueError):
        return None
    if np.isnan(score):
        return None
    if score < _QUALITY_BAD_THRESHOLD:
        return _QUALITY_BAD
    if score < _QUALITY_WARN_THRESHOLD:
        return _QUALITY_WARN
    return _QUALITY_GOOD


def _cell_value(value: Any) -> Any:
    """openpyxl rejects numpy scalars; NaN/NaT become empty cells."""
    if value is None:
        return None
    if isinstance(value, np.generic):
        value = value.item()
    if isinstance(value, float) and np.isnan(value):
        return None
    if pd.api.types.is_scalar(value) and pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if not isinstance(value, (bool, int, float, str, bytes)) and value is not None:
        return str(value)
    return value


def _write_sheet(worksheet: Worksheet, frame: pd.DataFrame) -> None:
    columns = [str(column) for column in frame.columns] or ["(empty)"]
    for column_index, column_name in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=column_name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    number_formats: list[str | None] = []
    quality_flags: list[bool] = []
    for column_name in columns:
        if column_name in frame.columns:
            series = frame[column_name]
        else:
            series = pd.Series(dtype=object)
        number_formats.append(_number_format_for(column_name, series))
        quality_flags.append(_is_quality_column(column_name, series))

    for row_offset, row in enumerate(frame.itertuples(index=False, name=None), start=2):
        for column_index, raw in enumerate(row, start=1):
            cell = worksheet.cell(row=row_offset, column=column_index, value=_cell_value(raw))
            fmt = number_formats[column_index - 1]
            if fmt is not None and isinstance(cell.value, (int, float)):
                cell.number_format = fmt
            if quality_flags[column_index - 1]:
                fill = _quality_fill(cell.value)
                if fill is not None:
                    cell.fill = fill

    last_column = get_column_letter(len(columns))
    last_row = max(2, len(frame) + 1)
    worksheet.auto_filter.ref = f"A1:{last_column}{last_row}"
    worksheet.freeze_panes = "A2"

    for column_index, column_name in enumerate(columns, start=1):
        if column_name in frame.columns and not frame.empty:
            content_width = int(frame[column_name].astype(str).str.len().max())
        else:
            content_width = 0
        width = min(48, max(10, len(column_name) + 2, content_width + 2))
        worksheet.column_dimensions[get_column_letter(column_index)].width = width


def export_excel_workbook(path: Path, sheets: dict[str, pd.DataFrame]) -> dict[str, Any]:
    """Write the 18-sheet match report workbook atomically, then re-validate it.

    ``sheets`` must contain exactly the sheet names in EXCEL_SHEET_NAMES; sheet
    order in the workbook always follows EXCEL_SHEET_NAMES regardless of the
    dict's insertion order.
    """
    path = Path(path)
    provided = set(sheets)
    expected = set(EXCEL_SHEET_NAMES)
    missing = sorted(expected - provided)
    unexpected = sorted(provided - expected)
    if missing or unexpected:
        raise ValueError(
            f"Workbook sheets mismatch; missing={missing} unexpected={unexpected}"
        )

    path.parent.mkdir(parents=True, exist_ok=True)
    # openpyxl only reopens files with a recognized extension, so keep .xlsx.
    fd, tmp_name = tempfile.mkstemp(dir=path.parent, prefix=path.name, suffix=".part.xlsx")
    os.close(fd)
    tmp_path = Path(tmp_name)
    try:
        workbook = Workbook()
        workbook.remove(workbook.active)
        for sheet_name in EXCEL_SHEET_NAMES:
            worksheet = workbook.create_sheet(title=sheet_name)
            _write_sheet(worksheet, sheets[sheet_name])
        workbook.save(tmp_path)
        workbook.close()
        report = validate_excel_workbook(tmp_path, expected_rows={
            name: len(frame) for name, frame in sheets.items()
        })
        os.replace(tmp_path, path)
    finally:
        tmp_path.unlink(missing_ok=True)
    report["path"] = str(path)
    return report


def validate_excel_workbook(
    path: Path, expected_rows: dict[str, int] | None = None
) -> dict[str, Any]:
    """Reopen the workbook with openpyxl and verify structure and formatting."""
    workbook = load_workbook(Path(path), read_only=False)
    try:
        if tuple(workbook.sheetnames) != EXCEL_SHEET_NAMES:
            raise RuntimeError(
                f"Sheet names mismatch in {path}: {workbook.sheetnames}"
            )
        rows_per_sheet: dict[str, int] = {}
        for sheet_name in EXCEL_SHEET_NAMES:
            worksheet = workbook[sheet_name]
            if worksheet.freeze_panes != "A2":
                raise RuntimeError(f"Missing freeze panes on '{sheet_name}' in {path}")
            if not worksheet.auto_filter.ref:
                raise RuntimeError(f"Missing auto filter on '{sheet_name}' in {path}")
            data_rows = max(0, (worksheet.max_row or 1) - 1)
            rows_per_sheet[sheet_name] = data_rows
            if expected_rows is not None and expected_rows[sheet_name] != data_rows:
                raise RuntimeError(
                    f"Row count mismatch on '{sheet_name}' in {path}: "
                    f"expected {expected_rows[sheet_name]}, found {data_rows}"
                )
    finally:
        workbook.close()
    return {
        "path": str(path),
        "sheets": list(EXCEL_SHEET_NAMES),
        "rows_per_sheet": rows_per_sheet,
        "validated": True,
    }
