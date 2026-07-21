from __future__ import annotations

import json
from pathlib import Path

import cv2
import numpy as np
import pandas as pd
import pyarrow as pa
import pytest
from openpyxl import load_workbook

from football_analytics.export.csv_exporter import export_csv
from football_analytics.export.excel_exporter import (
    EXCEL_SHEET_NAMES,
    export_excel_workbook,
    validate_excel_workbook,
)
from football_analytics.export.json_exporter import export_json
from football_analytics.export.parquet_exporter import export_parquet
from football_analytics.export.tactical_map_exporter import export_tactical_map_video
from football_analytics.export.video_exporter import (
    export_annotated_video,
    export_review_grid_video,
    validate_video_export,
)


def _write_synthetic_video(
    path: Path, frames: int = 24, fps: float = 12.0, size: tuple[int, int] = (320, 240)
) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    writer = cv2.VideoWriter(str(path), cv2.VideoWriter_fourcc(*"mp4v"), fps, size)
    assert writer.isOpened()
    for frame_id in range(frames):
        image = np.full((size[1], size[0], 3), (frame_id * 9 % 255, 60, 120), dtype=np.uint8)
        cv2.putText(
            image, f"f{frame_id}", (10, 40), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (255, 255, 255), 2
        )
        writer.write(image)
    writer.release()
    return path


# ---------------------------------------------------------------- JSON


def test_json_export_round_trip_and_atomicity(tmp_path):
    payload = {
        "match_id": "match-01",
        "frames": np.int64(1200),
        "quality": np.float32(0.875),
        "invalid_metric": float("nan"),
        "cameras": [Path("cam1.mp4"), Path("cam2.mp4")],
        "flags": np.array([1, 2, 3]),
    }
    target = tmp_path / "reports" / "summary.json"
    report = export_json(target, payload)

    assert report["validated"] is True
    data = json.loads(target.read_text(encoding="utf-8"))
    assert data["frames"] == 1200
    assert data["quality"] == pytest.approx(0.875)
    assert data["invalid_metric"] is None
    assert data["cameras"] == ["cam1.mp4", "cam2.mp4"]
    assert data["flags"] == [1, 2, 3]
    assert not list(target.parent.glob("*.part")), "temporary file leaked"


# ---------------------------------------------------------------- CSV


def test_csv_export_round_trip_and_no_partial_files(tmp_path):
    frame = pd.DataFrame(
        {
            "player_id": ["p1", "p2", "p3"],
            "distance_m": [10234.5, 9800.25, 11002.0],
            "team_id": ["team_0", "team_0", "team_1"],
        }
    )
    target = tmp_path / "player_summary.csv"
    report = export_csv(target, frame)

    assert report["rows"] == 3
    read_back = pd.read_csv(target)
    assert list(read_back.columns) == ["player_id", "distance_m", "team_id"]
    assert read_back["distance_m"].tolist() == [10234.5, 9800.25, 11002.0]
    assert not list(tmp_path.glob("*.part"))


# ---------------------------------------------------------------- Parquet


def test_parquet_export_dataframe_round_trip(tmp_path):
    frame = pd.DataFrame(
        {"frame_id": [0, 1, 2], "confidence": [0.9, 0.4, 0.7], "valid": [True, False, True]}
    )
    target = tmp_path / "tracks.parquet"
    report = export_parquet(target, frame)

    assert report["rows"] == 3
    read_back = pd.read_parquet(target)
    pd.testing.assert_frame_equal(read_back, frame)


def test_parquet_export_typed_empty_rows_with_schema(tmp_path):
    schema = pa.schema(
        [pa.field("frame_id", pa.int64()), pa.field("confidence", pa.float64())]
    )
    target = tmp_path / "empty.parquet"
    report = export_parquet(target, rows=[], schema=schema)

    assert report["rows"] == 0
    read_back = pa.parquet.read_table(target)
    assert read_back.schema.equals(schema, check_metadata=False)


def test_parquet_export_rejects_ambiguous_inputs(tmp_path):
    with pytest.raises(ValueError):
        export_parquet(tmp_path / "x.parquet")
    with pytest.raises(ValueError):
        export_parquet(tmp_path / "x.parquet", rows=[{"a": 1}])


# ---------------------------------------------------------------- XLSX


def _workbook_sheets() -> dict[str, pd.DataFrame]:
    sheets = {name: pd.DataFrame({"note": [f"{name} placeholder"]}) for name in EXCEL_SHEET_NAMES}
    sheets["Match Summary"] = pd.DataFrame(
        {"match_id": ["match-01"], "duration_minutes": [94.5], "total_goals": [3]}
    )
    sheets["Player Summary"] = pd.DataFrame(
        {
            "player_id": ["p1", "p2", "p3"],
            "distance_km": [10.4, 9.1, 11.2],
            "identity_confidence": [0.92, 0.61, 0.35],
        }
    )
    sheets["Visibility Quality"] = pd.DataFrame(
        {"chunk_id": [0, 1], "visibility_score": [0.88, 0.42]}
    )
    sheets["Camera Coverage"] = pd.DataFrame(
        {"camera": ["cam1", "cam2"], "coverage_pct": [0.97, 0.55]}
    )
    sheets["Errors and Warnings"] = pd.DataFrame(columns=["severity", "message"])
    return sheets


def test_excel_workbook_has_exactly_18_sheets_with_formatting(tmp_path):
    target = tmp_path / "match_report.xlsx"
    report = export_excel_workbook(target, _workbook_sheets())

    assert report["validated"] is True
    workbook = load_workbook(target)
    assert tuple(workbook.sheetnames) == EXCEL_SHEET_NAMES
    assert len(workbook.sheetnames) == 18

    for sheet_name in EXCEL_SHEET_NAMES:
        worksheet = workbook[sheet_name]
        assert worksheet.freeze_panes == "A2", sheet_name
        assert worksheet.auto_filter.ref, sheet_name
        header = worksheet.cell(row=1, column=1)
        assert header.font.bold
        assert worksheet.column_dimensions["A"].width >= 10

    players = workbook["Player Summary"]
    confidence_column = None
    for column_index in range(1, players.max_column + 1):
        if players.cell(row=1, column=column_index).value == "identity_confidence":
            confidence_column = column_index
    assert confidence_column is not None
    fills = {
        round(players.cell(row=row, column=confidence_column).value, 2): players.cell(
            row=row, column=confidence_column
        ).fill.start_color.rgb
        for row in range(2, 5)
    }
    assert fills[0.92].endswith("C6EFCE")  # good: green
    assert fills[0.61].endswith("FFEB9C")  # warn: yellow
    assert fills[0.35].endswith("FFC7CE")  # bad: red
    assert (
        players.cell(row=2, column=confidence_column).number_format == "0.000"
    )
    coverage = workbook["Camera Coverage"]
    assert coverage.cell(row=2, column=2).number_format == "0.0%"
    workbook.close()


def test_excel_workbook_rejects_wrong_sheet_sets(tmp_path):
    sheets = _workbook_sheets()
    removed = sheets.pop("Officials")
    with pytest.raises(ValueError, match="missing"):
        export_excel_workbook(tmp_path / "bad.xlsx", sheets)
    sheets["Officials"] = removed
    sheets["Extra Sheet"] = pd.DataFrame()
    with pytest.raises(ValueError, match="unexpected"):
        export_excel_workbook(tmp_path / "bad.xlsx", sheets)
    assert not list(tmp_path.glob("*.xlsx")), "no workbook may be written on failure"


def test_excel_validation_helper_reopens_workbook(tmp_path):
    target = tmp_path / "match_report.xlsx"
    sheets = _workbook_sheets()
    export_excel_workbook(target, sheets)
    report = validate_excel_workbook(
        target, expected_rows={name: len(frame) for name, frame in sheets.items()}
    )
    assert report["rows_per_sheet"]["Player Summary"] == 3
    assert report["rows_per_sheet"]["Errors and Warnings"] == 0


# ---------------------------------------------------------------- MP4


def test_annotated_video_export_streams_and_validates(tmp_path):
    source = _write_synthetic_video(tmp_path / "input.mp4", frames=24, fps=12.0)
    seen_frames: list[int] = []

    def annotate(image: np.ndarray, frame_id: int, timestamp_ms: float) -> None:
        seen_frames.append(frame_id)
        cv2.circle(image, (50, 50), 10, (0, 255, 0), -1)

    report = export_annotated_video(source, tmp_path / "annotated.mp4", annotate=annotate)

    assert report["frames"] == 24
    assert seen_frames == list(range(24))
    assert report["validation"]["validated"] is True
    assert report["validation"]["ffprobe"]["width"] == 320


@pytest.mark.parametrize("camera_count,expected_size", [(2, (640, 180)), (4, (640, 360))])
def test_review_grid_export_for_2_and_4_cameras(tmp_path, camera_count, expected_size):
    cameras = [
        _write_synthetic_video(tmp_path / f"cam{index}.mp4", frames=20, fps=10.0)
        for index in range(camera_count)
    ]
    report = export_review_grid_video(
        cameras, tmp_path / f"grid{camera_count}.mp4", tile_size=(320, 180)
    )

    assert report["frames"] == 20
    assert (report["width"], report["height"]) == expected_size
    assert report["cameras"] == [f"CAM {index + 1}" for index in range(camera_count)]
    assert report["validation"]["validated"] is True


def test_review_grid_synchronizes_mismatched_frame_rates(tmp_path):
    fast = _write_synthetic_video(tmp_path / "fast.mp4", frames=40, fps=20.0)
    slow = _write_synthetic_video(tmp_path / "slow.mp4", frames=20, fps=10.0)
    report = export_review_grid_video([fast, slow], tmp_path / "grid.mp4", fps=20.0)

    # Both cameras cover 2 seconds; output clock at 20 fps yields ~40 frames.
    assert abs(report["frames"] - 40) <= 1
    assert report["fps"] == 20.0


def test_review_grid_rejects_unsupported_camera_counts(tmp_path):
    source = _write_synthetic_video(tmp_path / "cam.mp4", frames=5)
    with pytest.raises(ValueError):
        export_review_grid_video([source], tmp_path / "grid.mp4")
    with pytest.raises(ValueError):
        export_review_grid_video([source, source, source], tmp_path / "grid.mp4")


def test_tactical_map_export_renders_positions_and_withheld_frames(tmp_path):
    positions = pd.DataFrame(
        {
            "frame_id": [0, 0, 1],
            "x_field": [52.5, 20.0, 80.0],
            "y_field": [34.0, 10.0, 50.0],
            "team_id": ["team_0", "team_1", "team_0"],
            "label": ["10", "7", "10"],
            "valid": [True, True, True],
        }
    )
    report = export_tactical_map_video(
        tmp_path / "tactical.mp4", positions, fps=10.0, frame_count=5
    )

    assert report["frames"] == 5
    assert report["valid_frames"] == 2  # frames 2-4 have no valid positions
    assert report["validation"]["ffprobe"]["width"] == 840
    assert report["validation"]["validated"] is True


def test_video_validation_helper_rejects_missing_file(tmp_path):
    with pytest.raises(RuntimeError):
        validate_video_export(tmp_path / "missing.mp4")
